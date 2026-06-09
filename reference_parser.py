#!/usr/bin/env python3
"""
Roster reference parser / offline validator
===========================================
Reads the daily roster .xlsx files (01JUN..04JUN style) and reports, for EVERY
team: headcount (working / off / sick / personal+vacation / OT people + OT hours)
and per-employee flight assignments with shift and flight times.

This is the ground-truth spec that RosterReader.gs mirrors in Google Apps Script.
The single most important rule discovered from the real data:

    *** The REMARK column is the source of truth for attendance. ***

A staff row may carry a shift code such as "X9" (00:00-09:00) yet REMARK says
"OFF" / "OFF (NO RQ OT)" / "OT OFF" — they are NOT working. Never infer
attendance from the shift code when REMARK is present.

Usage:  python3 reference_parser.py <file.xlsx> [TEAM]
"""
import openpyxl, re, sys


# ── cell helpers ────────────────────────────────────────────────────────────
def cv(v):
    """Clean a cell to a trimmed string; times -> HH:MM; strip trailing .0"""
    if v is None:
        return ''
    if hasattr(v, 'hour'):                       # datetime / time object
        return f"{v.hour:02d}:{v.minute:02d}" if (v.hour or v.minute) else ''
    s = str(v).strip()
    return s[:-2] if s.endswith('.0') else s


def up(v):
    return cv(v).upper()


SKIP = {'MANPOWER', 'ROSTER', 'SUMMARY', 'MASTER SMART SHIFT', 'SHIFTDB', 'CODE'}


# ── attendance classification (REMARK-driven) ──────────────────────────────
def classify(shift, remark):
    rm = cv(remark).upper().strip()
    sh = cv(shift).upper().strip()
    core = re.sub(r'\(.*?\)', '', rm).strip()    # drop "(NO RQ OT)" notes

    if core.startswith('SICK') or core in ('SL', 'MC') or sh in ('SICK', 'SL', 'MC'):
        return 'sick'
    if core.startswith('VAC') or core in ('BL', 'AL', 'VACATION'):
        return 'vac'
    if core.startswith('OT OFF') or core.startswith('OT-OFF'):
        return 'ot_off'
    if core.startswith('ONDUTY') or core.startswith('ON DUTY'):
        return 'working'
    if core.startswith('OFF') or core == 'X':
        return 'off'
    if core == '':                               # no REMARK -> fall back to shift
        if 'VAC' in sh or sh == 'BL':
            return 'vac'
        if sh in ('SL', 'SICK', 'MC'):
            return 'sick'
        if sh in ('', 'X', 'XX', 'OFF', '-') or sh.startswith('OFF'):
            return 'off'
        return 'working'
    return 'working'


# ── time / OT helpers ──────────────────────────────────────────────────────
def parse_time_pair(s):
    m = re.search(r'(\d{1,2})[:.]?(\d{2})', cv(s))
    return f"{m.group(1).zfill(2)}:{m.group(2)}" if m else ''


def hours_from_range(s):
    s = cv(s).replace('.', ':')
    m = re.match(r'^(\d{1,2}):?(\d{2})?\s*[-–]\s*(\d{1,2}):?(\d{2})?', s)
    if not m:
        return 0.0
    a = int(m.group(1)) * 60 + (int(m.group(2)) if m.group(2) else 0)
    b = int(m.group(3)) * 60 + (int(m.group(4)) if m.group(4) else 0)
    if b <= a:
        b += 1440
    return round((b - a) / 60, 1)


def _to_min(v):
    s = cv(v)
    m = re.match(r'^(\d{1,2})[:.](\d{2})', s)
    if m:
        return int(m.group(1)) * 60 + int(m.group(2))
    m = re.match(r'^(\d{2})(\d{2})$', s)
    return int(m.group(1)) * 60 + int(m.group(2)) if m else None


def _range_str(s):
    m = re.search(r'(\d{1,2}):?(\d{2})?\s*[-–]\s*(\d{1,2}):?(\d{2})?', cv(s))
    if m:
        return [int(m.group(1)) * 60 + (int(m.group(2)) if m.group(2) else 0),
                int(m.group(3)) * 60 + (int(m.group(4)) if m.group(4) else 0)]
    return [None, None]


def _range_cells(row, col):
    if col is None or col < 0 or col >= len(row):
        return [None, None]
    r = _range_str(row[col])
    if r[0] is not None:
        return r
    return [_to_min(row[col]), _to_min(row[col + 1]) if col + 1 < len(row) else None]


def ot_type(srng, orng, is_off):
    if is_off:
        return 'POST'
    si, so = srng
    oi, oo = orng
    if oi is None:
        return 'POST'
    if so is not None and si is not None and so <= si:
        so += 1440
    if oo is not None and oo <= oi:
        oo += 1440
    TOL = 30
    if si is not None and oo is not None and oo <= si + TOL:
        return 'PRE'
    if so is not None and oi >= so - TOL:
        return 'POST'
    if si is not None and oi < si:
        return 'PRE'
    return 'POST'


def ot_hours(v):
    """OT 'Total Hrs' cell -> decimal hours. Accepts 1.5 / '1:30' / '17-21'."""
    s = cv(v).upper()
    if not s or s in ('-', 'NO OT', 'VAC', 'X'):
        return 0.0
    m = re.match(r'^(\d{1,2}):(\d{2})(:\d{2})?$', s)        # duration H:MM[:SS]
    if m:
        h, mi = int(m.group(1)), int(m.group(2))
        return round(h + mi / 60, 1) if h <= 14 else 0.0   # >14 => clock time
    try:
        f = float(s)
        return f if 0 < f <= 14 else 0.0
    except ValueError:
        pass
    return hours_from_range(s)


# ── header detection (works for standard AND the TR NO/ID/NAME/TIME variant) ─
def find_header(rows):
    for r in range(min(8, len(rows))):
        u = [up(c) for c in rows[r]]
        if 'NAME' not in u:
            continue
        if not ({'ID', 'NO', 'NO.'} & set(u)):
            continue
        cm = {'hdr': r, 'name': u.index('NAME')}
        cm['id']     = u.index('ID') if 'ID' in u else (u.index('NO') if 'NO' in u else u.index('NO.'))
        cm['shift']  = u.index('SHIFT') if 'SHIFT' in u else -1
        cm['time']   = u.index('TIME') if 'TIME' in u else -1
        cm['pos']    = u.index('POSITION') if 'POSITION' in u else (u.index('POS.') if 'POS.' in u else -1)
        cm['remark'] = u.index('REMARK') if 'REMARK' in u else -1
        cm['resked'] = u.index('RE-SKED') if 'RE-SKED' in u else (u.index('RESKED') if 'RESKED' in u else -1)
        cm['ot']     = u.index('OT') if 'OT' in u else -1
        # OT total-hours column = a "Total Hrs" header within 3 cols after OT
        tothrs = [c for c, h in enumerate(u)
                  if h.replace('.', '').replace('  ', ' ').strip().startswith('TOTAL')]
        cm['ottot'] = next((c for c in tothrs if cm['ot'] >= 0 and 0 < c - cm['ot'] <= 3), -1)
        cm['flt'] = u.index('FLIGHT') + 1 if 'FLIGHT' in u else -1
        if cm['flt'] < 0:
            # บางวันชีต (เช่น WY) ไม่มีหัว "FLIGHT" — รหัสไฟลท์อยู่ในหัวตารางตรง ๆ
            # → หาคอลัมน์แรกหลังคอลัมน์ข้อมูลที่หัวเป็น "รหัสไฟลท์"
            after = max(cm['remark'], cm['ot'], cm['ottot'], cm['time'], cm['shift'], cm['name'], cm['id'])
            for c in range(after + 1, len(u)):
                if _is_flight_hdr(u[c]):
                    cm['flt'] = c
                    break
        return r, cm
    return None, None


def _is_flight_hdr(h):
    """หัวคอลัมน์ที่เป็น 'รหัสไฟลท์' (G9687/688, WY831/832, CA413, 6E1077, SQ726)."""
    return bool(re.search(r'(?:^|[\s/])(?:[A-Z]{1,3}\s?\d{2,4}|\d[A-Z]\d{2,4})', str(h or '')))


# ── parsers ────────────────────────────────────────────────────────────────
def parse_standard(rows, team):
    hi, cm = find_header(rows)
    if hi is None:
        return None

    flights, fltcols = {}, []
    if cm['flt'] >= 0:
        hdr = rows[hi]
        for c in range(cm['flt'], len(hdr)):
            nm = cv(hdr[c])
            if nm and not nm.startswith('=') and nm.upper() not in (
                    'STA / STD', 'OP / CL', 'REMARK', 'RE', 'OT', 'COUNTER',
                    'NIL', '-', 'N/A', 'NA'):    # NIL = placeholder "ไม่มีไฟลท์" — ไม่นับ
                fltcols.append((c, nm))
        sta = rows[hi + 1] if hi + 1 < len(rows) else []
        opn = rows[hi + 2] if hi + 2 < len(rows) else []
        spans = []
        for i, (c, nm) in enumerate(fltcols):
            c1 = fltcols[i + 1][0] if i + 1 < len(fltcols) else len(hdr)
            spans.append((c, c1, nm))
            st, oc = [], []
            for cc in range(c, c1):
                tv = parse_time_pair(sta[cc]) if cc < len(sta) else ''
                if tv:
                    st.append(tv)
                ov = parse_time_pair(opn[cc]) if cc < len(opn) else ''
                if ov:
                    oc.append(ov)
            flights[nm] = {'STA': st[0] if st else '', 'STD': st[1] if len(st) > 1 else '',
                           'OP': oc[0] if oc else '', 'CL': oc[1] if len(oc) > 1 else ''}
        fltcols = spans  # (c, c_end, name)

    recs, seen = [], set()
    for r in range(hi + 1, len(rows)):
        row = rows[r]
        idd = re.sub(r'\D', '', cv(row[cm['id']])) if cm['id'] < len(row) else ''
        if len(idd) < 6 and cm['id'] + 1 < len(row):          # WY has a seq col first
            raw_next = re.sub(r'\.0+$', '', cv(row[cm['id'] + 1]))
            if re.fullmatch(r'\d{6,8}', raw_next):            # only a PURE numeric id (not a flight code like PG251/252)
                idd = raw_next
        name = cv(row[cm['name']]) if cm['name'] < len(row) else ''
        if not name or not (6 <= len(idd) <= 8):
            continue
        if name.upper() in ('NAME', 'REMARK', 'SUPPORT', 'JAIDEE') or idd in seen:
            continue
        if up(row[cm['id']]).startswith('EX'):       # template "Ex. 212121" sample row
            continue
        seen.add(idd)

        shift  = cv(row[cm['shift']]) if 0 <= cm['shift'] < len(row) else ''
        timev  = cv(row[cm['time']]) if 0 <= cm['time'] < len(row) else ''
        remark = cv(row[cm['remark']]) if 0 <= cm['remark'] < len(row) else ''
        # OT ชั่วโมง: ปกติอยู่คอลัมน์ TOTAL; ถ้าไม่มี → ใช้ค่าในคอลัมน์ OT เอง (เช่น "14-20")
        ot_col = cm['ottot'] if cm['ottot'] >= 0 else cm['ot']
        otv    = cv(row[ot_col]) if 0 <= ot_col < len(row) else ''

        assigns = []
        for c, c1, nm in fltcols:
            tasks = [cv(row[cc]) for cc in range(c, c1) if cc < len(row) and cv(row[cc])]
            if tasks:
                assigns.append(dict(flight=nm, task='/'.join(tasks), **flights.get(nm, {})))
        # บางเทมเพลต (เช่น REV.01 TK) เขียนไฟลท์เป็นข้อความในคอลัมน์ "FLIGHT" (เช่น VJ808/OD543)
        # → เพิ่มรหัสไฟลท์ที่ยังไม่มี (กันนับซ้ำด้วยเลขไฟลท์)
        if cm['flt'] - 1 >= 0:
            label = cv(row[cm['flt'] - 1]) if cm['flt'] - 1 < len(row) else ''
            if label and not re.match(r'(OFF|VAC|SICK|SL|BL|X|ONDUTY|SUPPORT|PASSENGER|NIL)', label, re.I):
                nums = set()
                for a in assigns:
                    nums.update(re.findall(r'\d{2,4}', a['flight']))
                for code in re.findall(r'[A-Z0-9]{1,3}\s?\d{2,4}(?:\s?/\s?\d{2,4})?', label, re.I):
                    code = code.strip()
                    cn = re.findall(r'\d{2,4}', code)
                    if cn and not all(n in nums for n in cn) and _ac_is_flight(code):
                        assigns.append(dict(flight=code, task='', STA='', STD='', OP='', CL=''))
                        nums.update(cn)
        oth = ot_hours(otv)
        bkt = classify(shift or timev, remark)
        if bkt == 'off' and oth > 0:            # SHIFT=X แต่มี OT = ทำ OT วันหยุด
            bkt = 'ot_off'
        # เวลากะ: ปกติอยู่คอลัมน์ TIME; ถ้าไม่มี → อ่านช่วงจากคอลัมน์ SHIFT เอง (เช่น "09-17")
        srng = _range_cells(row, cm['time']) if cm['time'] >= 0 else _range_str(shift)
        re_time = ''
        if cm.get('resked', -1) >= 0:                         # Re-Sked overrides shift time
            rs = _range_cells(row, cm['resked'])
            if rs[0] is not None:
                srng = rs
        orng = _range_cells(row, cm['ot'])
        otype = ot_type(srng, orng, bkt == 'ot_off') if oth > 0 else None
        fmt = lambda m: ('%02d:%02d' % ((m // 60) % 24, m % 60)) if m is not None else ''
        srng_s = ('%s-%s' % (fmt(srng[0]), fmt(srng[1]))) if srng[0] is not None and srng[1] is not None else ''
        orng_s = ('%s-%s' % (fmt(orng[0]), fmt(orng[1]))) if oth > 0 and orng[0] is not None and orng[1] is not None else ''
        if cm.get('resked', -1) >= 0:
            rs2 = _range_cells(row, cm['resked'])
            if rs2[0] is not None:
                re_time = '%s-%s' % (fmt(rs2[0]), fmt(rs2[1])) if rs2[1] is not None else fmt(rs2[0])
        recs.append(dict(team=team, id=idd, name=name,
                         pos=cv(row[cm['pos']]) if cm['pos'] >= 0 else '', re=re_time,
                         shift=shift or timev, shift_time=srng_s or (shift or timev),
                         shift_start=srng[0] if srng[0] is not None else 99999,
                         bucket=bkt, ot=oth, ot_type=otype, ot_time=orng_s, assigns=assigns))
    return recs


def parse_porter(rows, team):
    recs = []
    for r in range(2, len(rows)):
        row = rows[r]
        for base in (0, 6):                       # two staff columns per row
            if base + 4 >= len(row):
                continue
            nm, sched, ot = cv(row[base]), cv(row[base + 3]), cv(row[base + 4])
            if (not nm or len(nm) < 2 or nm[0].isdigit()
                    or nm.upper() in ('NAME', '(INTER)', '(DOM)') or 'STBY' in nm.upper()):
                continue
            recs.append(dict(team=team, id='', name=nm, pos='PORTER', shift=sched,
                             bucket=classify(sched, ''), ot=ot_hours(ot), assigns=[]))
    return recs


def parse_admindoc(rows, team):
    recs = []
    for r in range(2, len(rows)):
        row = rows[r]
        nm = cv(row[0])
        sched = cv(row[1]) if len(row) > 1 else ''
        if not nm or len(nm) < 2 or nm.upper() in ('NAME', 'SCHEDULE'):
            continue
        flts = [cv(c) for c in row[2:] if cv(c)]
        recs.append(dict(team=team, id='', name=nm, pos='ADMINDOC', shift=sched,
                         bucket='off' if not sched or sched.upper() == 'OFF' else 'working',
                         ot=0.0, assigns=[dict(flight=f, task='') for f in flts]))
    return recs


def parse_crewsign(rows, team):
    recs, hi, seen = [], -1, set()
    for r in range(min(20, len(rows))):
        u = [up(c) for c in rows[r]]
        if 'STAFF NAME' in u or ('SHIFT' in u and 'REMARK' in u):
            hi = r
            break
    if hi < 0:
        return recs
    for r in range(hi + 1, len(rows)):
        row = rows[r]
        shift = cv(row[0])
        name = cv(row[1]) if len(row) > 1 else ''
        flt = cv(row[3]) if len(row) > 3 else ''
        if not name or len(name) < 2 or name.upper() in ('STAFF NAME', 'NAME'):
            continue
        key = re.sub(r'[\s.]+', '', name.upper())
        if key in seen:
            continue
        seen.add(key)
        actual = shift.split('/')[-1].strip() if '/' in shift else shift
        bkt = classify(actual, '') if shift else classify('', flt)
        recs.append(dict(team=team, id='', name=name, pos='CREWSIGN', shift=shift,
                         bucket=bkt, ot=0.0,
                         assigns=[dict(flight=flt, task='')] if (flt and flt.upper() != 'OFF') else []))
    return recs


def _su_name(raw):
    """Clean an SU cell to a real staff name, or None for codes/roles/porters."""
    n = re.sub(r'\s+(WK|TRN|EK|WY|QR|JQ|KC|ZF|FC|BOGO|PVT)\b.*$', '', str(raw or '').strip(), flags=re.I).strip()
    if not n or len(n) < 2 or n == '-':
        return None
    u = n.upper()
    if re.search(r'\d', u) or 'PORTER' in u:
        return None
    if u in ('SPVR', 'SOD', 'OB', 'ONBOARD', 'RF', 'CS', 'ARR', 'PSC', 'STBY',
             'SCAN', 'FILE', 'MONITOR', 'BRIEF', 'NIL', 'REMARK', 'GATE', 'AGENT', 'PREPARED'):
        return None
    return n


def parse_su(rows, team):
    """SU bespoke 3-section template: counter rotation + gate + job detail."""
    staff, info = {}, {}

    def get(raw):
        n = _su_name(raw)
        if not n:
            return None
        staff.setdefault(n, {'counter': [], 'flights': []})
        return n

    ci = ga = jb = -1
    for r in range(min(40, len(rows))):
        row = rows[r]
        c1 = up(row[1]) if len(row) > 1 else ''
        c2 = up(row[2]) if len(row) > 2 else ''
        c3 = up(row[3]) if len(row) > 3 else ''
        c5 = up(row[5]) if len(row) > 5 else ''
        if ci < 0 and c1 == 'FLT' and c2 in ('TIME', 'SCHEDULE'):
            ci = r
        elif ga < 0 and c1 == 'FLT' and 'GATE' in c3:
            ga = r
        elif jb < 0 and c1 == 'FLT' and 'SOD' in c5:
            jb = r

    if ci >= 0:                                              # counter rotation
        curflt = ''
        for r in range(ci + 1, len(rows)):
            row = rows[r]
            f = cv(row[1]) if len(row) > 1 else ''
            slot = cv(row[2]) if len(row) > 2 else ''
            if up(row[1]).startswith('ARRIVAL') or up(row[1]) == 'FLT':
                break
            if not slot:
                continue
            if f:
                curflt = f.replace('\n', ' ')
            for c in range(3, len(row)):
                for part in re.split(r'[,/]', cv(row[c])):
                    nm = get(part)
                    if nm:
                        staff[nm]['counter'].append({'flts': curflt, 'time': slot})

    if ga >= 0:                                              # gate per-flight
        groles = [cv(x) for x in rows[ga][3:]]
        for r in range(ga + 1, len(rows)):
            row = rows[r]
            flt = cv(row[1]) if len(row) > 1 else ''
            if not re.search(r'SU\d', flt, re.I):
                continue
            sta = cv(row[2]) if len(row) > 2 else ''
            info.setdefault(flt, {}).update(STA=sta.split('/')[0], STD=sta.split('/')[1] if '/' in sta else '')
            for c in range(3, len(row)):
                role = groles[c - 3] if c - 3 < len(groles) else 'GATE'
                for part in re.split(r'[,/]', cv(row[c])):
                    if up(part) == 'SPVR':
                        continue
                    nm = get(part)
                    if nm:
                        staff[nm]['flights'].append(dict(flight=flt, task=role, **info.get(flt, {})))

    if jb >= 0:                                              # job detail
        jroles = [cv(x) for x in rows[jb][5:]]
        for r in range(jb + 1, len(rows)):
            row = rows[r]
            flt = cv(row[1]) if len(row) > 1 else ''
            if not re.search(r'SU\d', flt, re.I):
                continue
            opcls = cv(row[4]) if len(row) > 4 else ''
            if '/' in opcls:
                p = opcls.split('/')
                info.setdefault(flt, {}).update(OP=p[0], CL=p[1])
            for c in range(5, len(row)):
                role = jroles[c - 5] if c - 5 < len(jroles) else ''
                for part in re.split(r'[,/]', cv(row[c])):
                    if up(part) == 'PORTER CS':
                        continue
                    nm = get(part)
                    if nm:
                        staff[nm]['flights'].append(dict(flight=flt, task=role, **info.get(flt, {})))

    recs = []
    for nm, d in staff.items():
        shift = ''
        if d['counter']:
            ts = [s['time'] for s in d['counter'] if re.search(r'[-–:]', s['time'])]
            if ts:
                shift = re.split(r'[-–]', ts[0])[0].strip() + '-' + re.split(r'[-–]', ts[-1])[-1].strip()
        assigns = list(d['flights'])
        if d['counter']:
            fset = sorted({x for s in d['counter'] for x in re.findall(r'SU\d+(?:/\d+)?', s['flts'], re.I)})
            assigns.insert(0, dict(flight='CHECK-IN COMMON', task=' '.join(fset),
                                   OP=d['counter'][0]['time'], CL=d['counter'][-1]['time']))
        recs.append(dict(team=team, id='', name=nm, pos='', shift=shift,
                         bucket='working' if (assigns or shift) else 'off', ot=0.0, assigns=assigns))
    return recs


def parse_sheet(ws, name):
    n = name.strip().upper()
    if any(s in n for s in SKIP):
        return None
    rows = list(ws.iter_rows(values_only=True))
    if 'PORTER' in n and 'CREW' in n:
        return parse_crewsign(rows, name)
    if n == 'PORTER':
        std = parse_standard(rows, name)
        return std if std else parse_porter(rows, name)
    if 'ADMIN' in n and 'DOC' in n:
        return parse_admindoc(rows, name)
    if n == 'SU' or n.startswith('SU '):
        std = parse_standard(rows, name)          # new SU template = standard table
        return std if std else parse_su(rows, name)  # else old counter-rotation grid
    return parse_standard(rows, name)


def pos_group(pos, team):
    """Map Position + team to an operational position group (exact, from file)."""
    t = str(team or '').upper()
    if 'CREW' in t:
        return 'Crewsign'
    if 'PORTER' in t:
        return 'Porter'
    if 'ADMIN' in t and 'DOC' in t:
        return 'AdminD'
    if 'GLOB' in t:
        return 'Globlex'
    c = re.sub(r'ACT\.?\s*', '', str(pos or '').upper()).strip()
    if 'DIRECTOR' in c:
        return 'DIR'
    if 'ASSIST' in c and 'MANAGER' in c:
        return 'Assist'
    if 'MANAGER' in c:
        return 'MGR'
    if 'SUP' in c or c == 'PSS':
        return 'PSS'
    if 'SNR' in c or 'SENIOR' in c:
        return 'SNR'
    if 'ADMIN' in c:
        return 'AdminD'
    if 'PORTER' in c:
        return 'Porter'
    return 'PSA'


def _rev_no(nm):
    m = re.search(r'REV\.?\s*0*(\d+)', nm, re.I)
    if m:
        return int(m.group(1))
    return 0 if re.search(r'REV', nm, re.I) else -1


def _team_base(nm):
    return re.sub(r'[\s._\-]+', '', re.sub(r'REV\.?\s*\d*', '', nm, flags=re.I)).upper()


def filter_rev(names):
    """When a team has several versions (AK, REV01 AK, REV02 AK) keep the latest REV."""
    maxrev = {}
    for nm in names:
        b = _team_base(nm); r = _rev_no(nm)
        if b not in maxrev or r > maxrev[b]:
            maxrev[b] = r
    out, taken = [], set()
    for nm in names:
        b = _team_base(nm)
        if _rev_no(nm) == maxrev[b] and b not in taken:
            taken.add(b); out.append(nm)
    return out


# ── report ─────────────────────────────────────────────────────────────────
def report(path, only_team=None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = filter_rev(wb.sheetnames)
    print("=" * 74)
    print("FILE:", path)
    hdr = f"{'TEAM':16}{'staff':>5}{'work':>5}{'off':>5}{'sick':>5}{'leave':>6}{'OTก่อน':>9}{'OTหลัง':>9}"
    print(hdr)
    tot = dict(staff=0, work=0, otoff=0, off=0, sick=0, leave=0, otp=0, oth=0.0,
               otpre=0, otpreh=0.0, otpost=0, otposth=0.0)
    for nm in names:
        recs = parse_sheet(wb[nm], nm)
        if recs is None:
            continue
        if not recs:
            print(f"{nm[:18]:18}{'  (no records parsed)':>40}")
            continue
        c = dict(work=0, otoff=0, off=0, sick=0, leave=0)
        for r in recs:
            b = r['bucket']
            c['work'] += b == 'working'
            c['otoff'] += b == 'ot_off'
            c['off'] += b == 'off'
            c['sick'] += b == 'sick'
            c['leave'] += b == 'vac'
        otp = sum(1 for r in recs if r['ot'] > 0)
        oth = round(sum(r['ot'] for r in recs), 1)
        otpre = sum(1 for r in recs if r['ot'] > 0 and r.get('ot_type') == 'PRE')
        otpreh = round(sum(r['ot'] for r in recs if r.get('ot_type') == 'PRE'), 1)
        otpost = sum(1 for r in recs if r['ot'] > 0 and r.get('ot_type') != 'PRE')
        otposth = round(sum(r['ot'] for r in recs if r['ot'] > 0 and r.get('ot_type') != 'PRE'), 1)
        pre_s = f"{otpre}({otpreh}h)" if otpre else '·'
        post_s = f"{otpost}({otposth}h)" if otpost else '·'
        print(f"{nm[:16]:16}{len(recs):>5}{c['work']:>5}{c['off']:>5}"
              f"{c['sick']:>5}{c['leave']:>6}{pre_s:>9}{post_s:>9}")
        for k in c:
            tot[k] += c[k]
        tot['staff'] += len(recs); tot['otp'] += otp; tot['oth'] += oth
        tot['otpre'] += otpre; tot['otpreh'] += otpreh; tot['otpost'] += otpost; tot['otposth'] += otposth

        if only_team and nm.strip().upper() == only_team.strip().upper():
            print(f"  --- per-employee flights/times for {nm} ---")
            for r in recs:
                if r['bucket'] not in ('working', 'ot_off'):
                    continue
                fl = "  ".join(f"{a['flight'].strip()}[{a['task'][:8]}]"
                               f"{a.get('OP','')}/{a.get('CL','')}" for a in r['assigns'])
                print(f"  {r['name'][:20]:20} {r['shift']:10} OT={r['ot']:>4}  "
                      f"flts={len(r['assigns']):>2}  {fl}")
    print("-" * 74)
    pre_s = f"{tot['otpre']}({round(tot['otpreh'],1)}h)"
    post_s = f"{tot['otpost']}({round(tot['otposth'],1)}h)"
    print(f"{'TOTAL':16}{tot['staff']:>5}{tot['work']:>5}{tot['off']:>5}"
          f"{tot['sick']:>5}{tot['leave']:>6}{pre_s:>9}{post_s:>9}")

    # by-position-group rollup (exact counts from the assignment file)
    pos = {}
    for nm in names:
        recs = parse_sheet(wb[nm], nm)
        if not recs:
            continue
        for r in recs:
            g = pos_group(r['pos'], nm)
            p = pos.setdefault(g, dict(staff=0, work=0, otoff=0, off=0, sick=0, leave=0, otp=0, oth=0.0))
            p['staff'] += 1
            p['work'] += r['bucket'] == 'working'
            p['otoff'] += r['bucket'] == 'ot_off'
            p['off'] += r['bucket'] == 'off'
            p['sick'] += r['bucket'] == 'sick'
            p['leave'] += r['bucket'] == 'vac'
            if r['ot'] > 0:
                p['otp'] += 1
                p['oth'] += r['ot']
    print("\nBY POSITION GROUP")
    print(f"{'POS':10}{'staff':>6}{'work':>6}{'otoff':>6}{'off':>5}{'sick':>5}{'leave':>6}{'otppl':>6}{'oth':>7}")
    for g in ['PSS', 'SNR', 'PSA', 'Globlex', 'AdminD', 'Porter', 'Crewsign', 'DIR', 'MGR', 'Assist']:
        if g not in pos:
            continue
        p = pos[g]
        print(f"{g:10}{p['staff']:>6}{p['work']:>6}{p['otoff']:>6}{p['off']:>5}"
              f"{p['sick']:>5}{p['leave']:>6}{p['otp']:>6}{round(p['oth'],1):>7}")
    wb.close()


# ── LL (baggage tracing) daily reader ───────────────────────────────────────
def ll_pos_group(pos):
    u = re.sub(r'ACT\.?\s*', '', str(pos or '').upper()).strip()
    if u.startswith('PSS') or 'SUPERVISOR' in u:
        return 'PSS'
    if u.startswith('SNR') or 'SENIOR' in u:
        return 'SNR'
    if 'TRAINEE' in u:
        return 'Trainee'
    if 'PORTER' in u:
        return 'Porter'
    if 'ADMIN' in u:
        return 'Admin'
    if u.startswith('PSA') or 'AGENT' in u:
        return 'PSA'
    return 'PSA'


def ll_classify(sched, remark):
    s = str(sched or '').upper().strip()
    rm = str(remark or '').upper()
    if s in ('SL', 'SICK', 'MC') or 'SICK' in rm:
        return 'sick'
    if s in ('VAC', 'BL', 'AL') or 'VAC' in s:
        return 'vac'
    if s in ('', 'OFF', 'X', 'XX') or s.startswith('OFF'):
        return 'off'
    return 'working'


def read_ll_tab(rows):
    recs, section, seen = [], '', set()
    for r in rows:
        c0 = cv(r[0]) if len(r) > 0 else ''
        if c0 and up(r[1] if len(r) > 1 else '') == 'NO' and up(r[2] if len(r) > 2 else '') == 'NAME':
            section = c0.replace('\n', ' ')
            continue
        no = cv(r[1]) if len(r) > 1 else ''
        name = cv(r[2]) if len(r) > 2 else ''
        pos = cv(r[3]) if len(r) > 3 else ''
        if not name or not pos or name.upper() == 'NAME' or not re.match(r'^\d+(\.\d+)?$', no):
            continue
        if name.upper() in seen:
            continue
        seen.add(name.upper())
        sched = cv(r[4]) if len(r) > 4 else ''
        remark = cv(r[6]) if len(r) > 6 else ''
        ot = cv(r[8]) if len(r) > 8 else ''
        recs.append(dict(section=section, name=name, pos=pos, posgroup=ll_pos_group(pos),
                         bucket=ll_classify(sched, remark), ot=ot_hours(ot)))
    return recs


def report_ll(path, tab=None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if tab is None:
        tab = next((s for s in wb.sheetnames if re.match(r'^\d{1,2}\s*[A-Z]{3}', s, re.I)), wb.sheetnames[0])
    recs = read_ll_tab(list(wb[tab].iter_rows(values_only=True)))
    wb.close()
    print("=" * 60)
    print(f"LL FILE: {path}  TAB: {tab}  staff: {len(recs)}")
    agg = {}
    for r in recs:
        p = agg.setdefault(r['posgroup'], dict(staff=0, work=0, off=0, sick=0, leave=0, otp=0, oth=0.0))
        p['staff'] += 1
        p['work'] += r['bucket'] == 'working'
        p['off'] += r['bucket'] == 'off'
        p['sick'] += r['bucket'] == 'sick'
        p['leave'] += r['bucket'] == 'vac'
        if r['ot'] > 0:
            p['otp'] += 1
            p['oth'] += r['ot']
    print(f"{'POSITION':10}{'staff':>6}{'work':>6}{'off':>5}{'sick':>5}{'leave':>6}{'otppl':>6}{'oth':>6}")
    tot = dict(staff=0, work=0, off=0, sick=0, leave=0, otp=0, oth=0.0)
    for g in ['PSS', 'SNR', 'PSA', 'Porter', 'Admin', 'Trainee']:
        if g not in agg:
            continue
        p = agg[g]
        print(f"{g:10}{p['staff']:>6}{p['work']:>6}{p['off']:>5}{p['sick']:>5}{p['leave']:>6}{p['otp']:>6}{round(p['oth'],1):>6}")
        for k in tot:
            tot[k] += p[k]
    print(f"{'TOTAL':10}{tot['staff']:>6}{tot['work']:>6}{tot['off']:>5}{tot['sick']:>5}{tot['leave']:>6}{tot['otp']:>6}{round(tot['oth'],1):>6}")


# ── assignment-quality check (mirror of AssignCheck.gs) ─────────────────────
AC_COVER_TOL = 45
AC_GAP_MIN = 180      # ช่วงว่างระหว่างไฟลท์ (split-duty dead time) ที่จะแจ้ง
AC_EDGE_MIN = 240     # ช่วงว่างก่อนไฟลท์แรก/หลังไฟลท์สุดท้าย (prep/standby) ที่จะแจ้ง
AC_IDLE_HRS = 7


def _ac_min(s):
    m = re.search(r'(\d{1,2}):(\d{2})', str(s if s is not None else ''))
    return int(m.group(1)) * 60 + int(m.group(2)) if m else None


def _ac_is_flight(name):
    """ไฟลท์จริง (มีรหัสไฟลท์) ที่ต้องเช็คการครอบคลุม — ไม่ใช่ pool/counter/common
    (CHECK-IN COMMON, LP MORNING/AFTERNOON, Counter G2/G11, Gate A1).
    รหัสไฟลท์จริงขึ้นต้นด้วยโค้ดสายการบิน; งานเคาน์เตอร์ขึ้นต้นด้วยคำอังกฤษ."""
    s = str(name or '').strip().upper()
    if not s:
        return False
    if re.match(r'(COUNTER|GATE|CHECK|ZONE|BELT|PIER|STBY|STAND|POOL|OFFICE|BRIEF|NIL|OFF\b|LP\s+(MORNING|AFTERNOON|NIGHT|DAY))', s):
        return False
    return bool(re.search(r'[A-Z]{1,3}\s*\d{2,4}', s))


def _ac_flight_win(a):
    # ช่วงที่ต้องอยู่ตามบทบาท: check-in→OP-CL, arrival→STA, gate→CL/STD, อื่น→min-max
    task = str(a.get('task') or '').upper()
    sta, op, cl, std = (_ac_min(a.get('STA', '')), _ac_min(a.get('OP', '')),
                        _ac_min(a.get('CL', '')), _ac_min(a.get('STD', '')))
    lo = hi = None
    if re.search(r'CT|CI|GK|CHECK|COUNTER|CTR', task) and op:
        lo, hi = op, (cl or std or op)
    elif re.search(r'ARR|MEET|\bAC\b|\bRF\b|ESCORT', task) and sta:
        lo, hi = sta, sta
    elif re.search(r'GATE|\bGA\b|\bGM\b|BOARD|\bGC\b|BIR|MAAS|PFD', task) and (cl or std):
        lo, hi = (cl or std), (std or cl)
    if lo is None:
        ts = [t for t in (sta, op, cl, std) if t]
        if not ts:
            return None
        lo, hi = min(ts), max(ts)
    if hi - lo > 14 * 60:
        hi -= 1440
    if hi < lo:
        lo, hi = hi, lo
    if hi - lo < 30:                 # ไฟลท์ที่มีเวลาจุดเดียว (เช่น PG STA=00:00 เหลือ STD)
        mid = (lo + hi) // 2         # → ให้เป็นบล็อกงาน ~60 นาที จะได้ตัด gap ถูก
        lo, hi = mid - 30, mid + 30
    return [lo, hi]


def _ac_duty(r):
    ss, se = _range_str(r.get('shift_time') or '')
    if ss is not None and se is not None and se <= ss:
        se += 1440
    oi, oo = _range_str(r.get('ot_time') or '')
    if oi is not None and oo is not None and oo <= oi:
        oo += 1440
    ds, de = ss, se
    if oi is not None:
        if ss is not None:
            while oi < ss - 720:
                oi += 1440
                oo += 1440
        ds = oi if ds is None else min(ds, oi)
        de = oo if de is None else max(de, oo)
    elif r.get('ot', 0) > 0 and ss is not None and r['bucket'] != 'ot_off':
        if r.get('ot_type') == 'PRE':
            ds = ss - round(r['ot'] * 60)
        else:
            de = se + round(r['ot'] * 60)
    return ss, se, ds, de


def _ac_fmt(m):
    return '%02d:%02d' % ((m // 60) % 24, m % 60) if m is not None else ''


def analyze_record(r):
    ss, se, ds, de = _ac_duty(r)
    # เชื่อถือได้เมื่อมีเวลากะจริง (ss) หรือเป็น OT OFF (ทำเฉพาะ OT วันหยุด).
    # ถ้าเป็นคนทำงานแต่กะเป็นรหัสไม่มีเวลา (เช่น NN0 กะดึก) → อย่าเอาช่วง OT มาตัดสิน coverage
    reliable = (ss is not None) or (r['bucket'] == 'ot_off' and ds is not None)
    a = dict(has=reliable and ds is not None and de is not None, shift='', duty='',
             flightN=0, coveredN=0, uncovered=[], gaps=[], otv='', issues=[], status='ok')
    a['shift'] = ('%s-%s' % (_ac_fmt(ss), _ac_fmt(se))) if ss is not None and se is not None else (r.get('shift') or '-')
    if a['has']:
        a['duty'] = '%s-%s' % (_ac_fmt(ds), _ac_fmt(de))
    wins = []
    for asg in r.get('assigns', []):
        if not asg.get('flight'):
            continue
        coverable = _ac_is_flight(asg['flight'])
        w = _ac_flight_win(asg)
        if not w:
            continue
        lo, hi = w
        if ds is not None and lo < ds - 720:
            lo += 1440
            hi += 1440
        wins.append((asg['flight'], lo, hi))     # ใช้ทุก task เพื่อหา gap (รวม pool)
        if not coverable:
            continue
        a['flightN'] += 1                         # นับเฉพาะไฟลท์จริงในการครอบคลุม
        if ds is not None and de is not None:
            if lo >= ds - AC_COVER_TOL and hi <= de + AC_COVER_TOL:
                a['coveredN'] += 1
            else:
                a['uncovered'].append('%s(%s-%s)' % (asg['flight'].strip(), _ac_fmt(lo), _ac_fmt(hi)))
    if a['has'] and wins:
        iv = sorted([[max(lo, ds), min(hi, de)] for _, lo, hi in wins if min(hi, de) > max(lo, ds)])
        merged = []
        for w in iv:
            if merged and w[0] <= merged[-1][1]:
                merged[-1][1] = max(merged[-1][1], w[1])
            else:
                merged.append(w)
        if merged:
            if merged[0][0] - ds >= AC_EDGE_MIN:                 # ก่อนไฟลท์แรก (edge)
                a['gaps'].append((ds, merged[0][0], 'edge'))
            for i in range(len(merged) - 1):                     # ระหว่างไฟลท์ (internal)
                if merged[i + 1][0] - merged[i][1] >= AC_GAP_MIN:
                    a['gaps'].append((merged[i][1], merged[i + 1][0], 'mid'))
            if de - merged[-1][1] >= AC_EDGE_MIN:                # หลังไฟลท์สุดท้าย (edge)
                a['gaps'].append((merged[-1][1], de, 'edge'))
    if a['uncovered']:
        a['status'] = 'bad'
        a['issues'].append('ไฟลท์นอกเวลางาน: ' + ', '.join(a['uncovered']))
        a['otv'] = 'OT ไม่พอ' if r.get('ot', 0) > 0 else 'ควรให้ OT/Re-Sked'
    elif r.get('ot', 0) > 0 and r['bucket'] != 'ot_off':
        just = a['flightN'] == 0
        for _, lo, hi in wins:
            if r.get('ot_type') == 'PRE' and ss is not None and lo < ss - AC_COVER_TOL:
                just = True
            if r.get('ot_type') != 'PRE' and se is not None and hi > se + AC_COVER_TOL:
                just = True
        if not just and a['flightN'] > 0:
            a['status'] = 'warn'
            a['otv'] = 'OT อาจเกินจำเป็น'
            a['issues'].append(a['otv'])
    if a['gaps']:
        if a['status'] == 'ok':
            a['status'] = 'warn'
        a['issues'].append('ช่วงว่าง ' + ', '.join(
            '%s-%s%s' % (_ac_fmt(x), _ac_fmt(y), '(ระหว่างไฟลท์)' if g == 'mid' else '')
            for x, y, g in a['gaps']))
    # ไม่มีไฟลท์เลย = นับเป็นข้อมูล (bench/standby/support) ไม่ flag เป็นปัญหา เพื่อไม่ให้ตารางรก
    a['noflt'] = a['flightN'] == 0 and r['bucket'] == 'working'
    return a


def report_assign(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = filter_rev(wb.sheetnames)
    print("=" * 74)
    print("ASSIGN CHECK:", path)
    s = dict(working=0, checked=0, bad=0, warn=0, otmuch=0, gap=0, nowin=0, noflt=0)
    flagged = []
    for nm in names:
        recs = parse_sheet(wb[nm], nm)
        if not recs:
            continue
        for r in recs:
            if r['bucket'] not in ('working', 'ot_off'):
                continue
            r['posgroup'] = pos_group(r.get('pos'), nm)
            s['working'] += 1
            a = analyze_record(r)
            if not a['has']:
                s['nowin'] += 1
                continue
            s['checked'] += 1
            if a['status'] == 'bad':
                s['bad'] += 1
            if a['status'] == 'warn':
                s['warn'] += 1
            if 'เกิน' in a['otv']:
                s['otmuch'] += 1
            if a['gaps']:
                s['gap'] += 1
            if a.get('noflt'):
                s['noflt'] += 1
            if a['status'] in ('bad', 'warn'):
                flagged.append((nm, r, a))
    order = {'bad': 0, 'warn': 1}
    flagged.sort(key=lambda x: (order[x[2]['status']], x[0]))
    emo = {'bad': 'X', 'warn': '!'}
    print("checked %d/%d  bad=%d warn=%d  OTเกิน=%d gap=%d  noflt=%d(support/standby)  (no-window %d skipped)"
          % (s['checked'], s['working'], s['bad'], s['warn'], s['otmuch'], s['gap'], s['noflt'], s['nowin']))
    print("-" * 74)
    for nm, r, a in flagged:
        ots = ('%sh %s' % (r['ot'], (r.get('ot_type') or 'OT'))) if r.get('ot', 0) > 0 else '-'
        print("%s %-10s %-18s กะ=%-12s OT=%-9s flts=%d/%d  %s"
              % (emo[a['status']], nm[:10], r['name'][:18], a['shift'], ots,
                 a['coveredN'], a['flightN'], ' · '.join(a['issues'])[:80]))
    wb.close()


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    if sys.argv[1] == '--ll':
        report_ll(sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else None)
    elif sys.argv[1] == '--assign':
        report_assign(sys.argv[2])
    else:
        report(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
